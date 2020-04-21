"Main command-line interface for program"

import datetime

import click
from openpyxl import Workbook, styles

import SRUMParse
import XLSXOutUtils
import ESEUtils
import SRUMColumns

ERROR_FONT = styles.Font(color=styles.colors.RED)


@click.command()
@click.option("--input", "-i", required=True, type=click.Path(exists=True,
                                                              dir_okay=False,
                                                              writable=False))

@click.option("--output", "-o", required=True, type=click.Path(exists=False,
                                                               dir_okay=False,
                                                               writable=True))

# Not currently implemented
# click.option("--include-registry", "-r", required=False, type=click.Path(exists=True,
#                                                                          file_okay=False))

@click.option("--omit-processed", "-p", is_flag=True, default=False)

@click.option("--only-processed", "-P", is_flag=True, default=False)

@click.pass_context
def export_xlsx(ctx, input, output, include_registry, omit_processed, only_processed):
    "Parse and export SRUM data from the --input provided ESE database file " \
    "to an .xlsx file in the --output directory, optionally including the SRUM " \
    "entries found in the registry in the folder provided by --include-registry."

    source_file = open(input, "rb")

    out_workbook = Workbook()

    if not only_processed:
        if include_registry is None:
            parser = SRUMParse.SRUMParser(source_file)
        else:
            parser = SRUMParse.SRUMParser(source_file, registryFolder=include_registry)
            
        for table in parser.raw_tables:
            print("Converting table", table.name)
            sheet_name = SRUMParse.short_table_name(table.name, False)
            worksheet = out_workbook.create_sheet(sheet_name)

            worksheet.append(["Full Table Name:", table.name])

            worksheet.append([col.name for col in table.columns])

            for row in parser.table_rows(table):
                nrow = [XLSXOutUtils.value_to_safe_string(col) for col in row]
                try:
                    worksheet.append(nrow)
                except Exception as e:
                    print("Encountered exception while adding a row to the sheet,\
                           this means that the row contains invalid characters.")
                    raise

    if not omit_processed:
        parser = SRUMParse.SRUMParser(source_file)

        SruDbIdMapTable = [table for table in parser.raw_tables if table.name=="SruDbIdMapTable"][0]
        SruDbIdMap = {}

        # Build a dict of items in the SruDbIdMapTable table, so values such as AppIds and UserIds
        # can be looked up without having to iterate over the entire table each time
        print("Building SruDbIdMapTable")
        SruDbIdStartTime = datetime.datetime.now()
        for row in parser.table_rows(SruDbIdMapTable):
            
            value_type = row[0]
            if value_type == 0 or row[2] == None:
                value = None
            elif value_type in [1, 2]: # UTF-16 string
                value = row[2].decode("utf-16")
            elif value_type == 3: # Windows SID
                try:
                    value = SRUMParse.SID_bytes_to_string(row[2])
                except Exception as e:
                    print(row)
                    raise e
            else:
                value = repr(row[2])
            SruDbIdMap[row[1]] = value

        print("Built SruDbIdMapTable in ", datetime.datetime.now()-SruDbIdStartTime)

        # --------------------------------
        # -------- PARSING TABLES --------
        # --------------------------------

        parser = SRUMParse.SRUMParser(source_file)

        for table in parser.raw_tables:
            
            print("Processing table", table.name)
            if table.name in ["{973F5D5C-1D90-4944-BE8E-24B94231A174}",
                              "{D10CA2FE-6FCF-4F6D-848E-B2E99266FA89}",
                              "{DD6636C4-8929-4683-974E-22C046A43763}",
                              "{D10CA2FE-6FCF-4F6D-848E-B2E99266FA86}",
                              "{5C8CF1C7-7257-4F13-B223-970EF5939312}",
                              "{FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}"]: # Is this table handled?
                column_names = []
                worksheet = out_workbook.create_sheet(SRUMParse.short_table_name(table.name, True))
                worksheet.append(["Original Table Name:", table.name])
                # Setup

                # --- Network Usage Data Monitor {973F5D5C-1D90-4944-BE8E-24B94231A174} ---
                if table.name == "{973F5D5C-1D90-4944-BE8E-24B94231A174}":
                    column_inserts = [["AppId", "P_AppId"],
                                      ["UserId", "P_UserId"],
                                      ["InterfaceLuid", "P_InterfaceLuid_IfType"],
                                      ["InterfaceLuid", "P_InterfaceLuid_NetLuidIndex"],
                                      ["BytesSent", "P_BytesSent"],
                                      ["BytesRecvd", "P_BytesRecvd"]]

                # --- Application Resource Usage {D10CA2FE-6FCF-4F6D-848E-B2E99266FA89} ---
                elif table.name == "{D10CA2FE-6FCF-4F6D-848E-B2E99266FA89}":
                    column_inserts = [["AppId", "P_AppId"],
                                      ["UserId", "P_UserId"],
                                      ["ForegroundBytesRead", "P_ForegroundBytesRead"],
                                      ["ForegroundBytesWritten", "P_ForegroundBytesWritten"],
                                      ["BackgroundBytesRead", "P_BackgroundBytesRead"],
                                      ["BackgroundBytesWritten", "P_BackgroundBytesWritten"]]  

                # --- Network Connectivity Usage Monitor {DD6636C4-8929-4683-974E-22C046A43763} ---
                elif table.name == "{DD6636C4-8929-4683-974E-22C046A43763}":
                    column_inserts = [["AppId", "P_AppId"],
                                      ["UserId", "P_UserId"],
                                      ["InterfaceLuid", "P_InterfaceLuid_IfType"],
                                      ["InterfaceLuid", "P_InterfaceLuid_NetLuidIndex"],
                                      ["ConnectedTime", "P_ConnectedTime"],
                                      ["ConnectStartTime", "P_ConnectStartTime"]]       

                # --- Push Notifications {D10CA2FE-6FCF-4F6D-848E-B2E99266FA86}  ---
                elif table.name == "{D10CA2FE-6FCF-4F6D-848E-B2E99266FA86}":
                    column_inserts = [["AppId", "P_AppId"],
                                      ["UserId", "P_UserId"],
                                      ["PayloadSize", "P_PayloadSize"],
                                      ["NotificationType", "P_NotificationType"]]

                # --- Energy Estimator {5C8CF1C7-7257-4F13-B223-970EF5939312}  ---
                elif table.name == "{5C8CF1C7-7257-4F13-B223-970EF5939312}":
                    column_inserts = [["AppId", "P_AppId"],
                                      ["UserId", "P_UserId"],
                                      ["EndTime", "P_EndTime"],
                                      ["DurationMS", "P_DurationMS"],
                                      ["SpanMS", "P_SpanMS"],
                                      ["NetworkBytesRaw", "P_NetworkBytesRaw"],
                                      ["MBBBytesRaw", "P_MBBBytesRaw"]]       

                # --- Energy Usage {FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}  ---
                elif table.name == "{FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}":
                    column_inserts = [["AppId", "P_AppId"],
                                      ["UserId", "P_UserId"],
                                      ["EventTimestamp", "P_EventTimestamp"]]

                # --- Long-Term Energy Usage {FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}LT  ---
                elif table.name == "{FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}":
                    column_inserts = [["AppId", "P_AppId"],
                                      ["UserId", "P_UserId"],
                                      ["ActiveAcTime", "P_ActiveAcTime"],
                                      ["CsAcTime", "P_CsAcTime"],
                                      ["ActiveDcTime", "P_ActiveDcTime"],
                                      ["CsDcTime", "P_CsDcTime"],
                                      ["ActiveDischargeTime", "P_ActiveDischargeTime"],
                                      ["CsDischargeTime", "P_CsDischargeTime"]]
                                    

                column_names = [col.name for col in table.columns]

                for insert in column_inserts:
                    column_names.insert(column_names.index(insert[0])+1, insert[1])

                worksheet.append(column_names)

                # Processing
                # Creates a dict of column names and their values, so that columns can be inserted
                # without needing to do complex shuffling. This dict is then converted back into
                # a list before being added to the worksheet.
                for row, raw_row in zip(parser.table_rows(table), parser.raw_table_rows(table)):
                    out_dict = {k:v for (k, v) in zip([col.name for col in table.columns], row)}

                    # --- Network Usage Data Monitor {973F5D5C-1D90-4944-BE8E-24B94231A174} ---
                    if table.name == "{973F5D5C-1D90-4944-BE8E-24B94231A174}":
                        luid_parsed = SRUMParse.parse_interface_luid(parser.row_element_by_column_name(raw_row, "InterfaceLuid", table))
                        for insert in column_inserts:
                            if insert[1] == "P_AppId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "AppId", table)]
                                #value = parser.row_element_by_column_name(row, "AppId", table)
                            elif insert[1] == "P_UserId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "UserId", table)]
                                #value = parser.row_element_by_column_name(row, "UserId", table)
                            elif insert[1] == "P_InterfaceLuid_IfType":
                                value = luid_parsed["iftype"]
                            elif insert[1] == "P_InterfaceLuid_NetLuidIndex":
                                value = luid_parsed["netluid_index"]
                            elif insert[1] == "P_BytesSent":
                                value = XLSXOutUtils.num_bytes_display(parser.row_element_by_column_name(row, "BytesSent", table))
                            elif insert[1] == "P_BytesRecvd":
                                value = XLSXOutUtils.num_bytes_display(parser.row_element_by_column_name(row, "BytesRecvd", table))
                            else:
                                value = "Could not parse value"
                            out_dict[insert[1]] = XLSXOutUtils.value_to_safe_string(value)
                    # --- Application Resource Usage {D10CA2FE-6FCF-4F6D-848E-B2E99266FA89} ---
                    elif table.name == "{D10CA2FE-6FCF-4F6D-848E-B2E99266FA89}":
                        for insert in column_inserts:
                            if insert[1] == "P_AppId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "AppId", table)]
                                #value = parser.row_element_by_column_name(row, "AppId", table)
                            elif insert[1] == "P_UserId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "UserId", table)]
                                #value = parser.row_element_by_column_name(row, "UserId", table)
                            elif insert[1] == "P_ForegroundBytesRead":
                                value = XLSXOutUtils.num_bytes_display(parser.row_element_by_column_name(row, "ForegroundBytesRead", table))
                            elif insert[1] == "P_ForegroundBytesWritten":
                                value = XLSXOutUtils.num_bytes_display(parser.row_element_by_column_name(row, "ForegroundBytesWritten", table))
                            elif insert[1] == "P_BackgroundBytesRead":
                                value = XLSXOutUtils.num_bytes_display(parser.row_element_by_column_name(row, "BackgroundBytesRead", table))
                            elif insert[1] == "P_BackgroundBytesWritten":
                                value = XLSXOutUtils.num_bytes_display(parser.row_element_by_column_name(row, "BackgroundBytesWritten", table))
                            else:
                                value = "Could not parse value"
                            out_dict[insert[1]] = XLSXOutUtils.value_to_safe_string(value)
                    # --- Network Connectivity Usage Monitor {DD6636C4-8929-4683-974E-22C046A43763} ---
                    elif table.name == "{DD6636C4-8929-4683-974E-22C046A43763}":
                        luid_parsed = SRUMParse.parse_interface_luid(parser.row_element_by_column_name(raw_row, "InterfaceLuid", table))
                        for insert in column_inserts:
                            if insert[1] == "P_AppId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "AppId", table)]
                                #value = parser.row_element_by_column_name(row, "AppId", table)
                            elif insert[1] == "P_UserId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "UserId", table)]
                                #value = parser.row_element_by_column_name(row, "UserId", table)
                            elif insert[1] == "P_InterfaceLuid_IfType":
                                value = luid_parsed["iftype"]
                            elif insert[1] == "P_InterfaceLuid_NetLuidIndex":
                                value = luid_parsed["netluid_index"]
                            elif insert[1] == "P_ConnectedTime":
                                value = XLSXOutUtils.num_secs_display(parser.row_element_by_column_name(row, "ConnectedTime", table))
                            elif insert[1] == "P_ConnectStartTime":
                                value = ESEUtils.parse_filetime(parser.row_element_by_column_name(row, "ConnectStartTime", table))
                            else:
                                value = "Could not parse value"
                            out_dict[insert[1]] = XLSXOutUtils.value_to_safe_string(value)

                    # --- Push Notifications {D10CA2FE-6FCF-4F6D-848E-B2E99266FA86} ---
                    elif table.name == "{D10CA2FE-6FCF-4F6D-848E-B2E99266FA86}":
                        for insert in column_inserts:
                            if insert[1] == "P_AppId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "AppId", table)]
                                #value = parser.row_element_by_column_name(row, "AppId", table)
                            elif insert[1] == "P_UserId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "UserId", table)]
                                #value = parser.row_element_by_column_name(row, "UserId", table)
                            elif insert[1] == "P_PayloadSize":
                                value = XLSXOutUtils.num_bytes_display(parser.row_element_by_column_name(row, "PayloadSize", table))
                            elif insert[1] == "P_NotificationType":
                                value = SRUMColumns.parse_notification_type(parser.row_element_by_column_name(row, "NotificationType", table))
                            else:
                                value = "Could not parse value"
                            out_dict[insert[1]] = XLSXOutUtils.value_to_safe_string(value)

                    # --- Energy Estimator {5C8CF1C7-7257-4F13-B223-970EF5939312}  ---
                    elif table.name == "{5C8CF1C7-7257-4F13-B223-970EF5939312}":
                        for insert in column_inserts:
                            if insert[1] == "P_AppId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "AppId", table)]
                                #value = parser.row_element_by_column_name(row, "AppId", table)
                            elif insert[1] == "P_UserId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "UserId", table)]
                                #value = parser.row_element_by_column_name(row, "UserId", table)
                            elif insert[1] == "P_EndTime":
                                value = ESEUtils.parse_filetime(parser.row_element_by_column_name(row, "EndTime", table))
                            elif insert[1] == "P_DurationMS":
                                value = XLSXOutUtils.num_secs_display(parser.row_element_by_column_name(row, "DurationMS", table)/1000)
                            elif insert[1] == "P_SpanMS":
                                value = XLSXOutUtils.num_secs_display(parser.row_element_by_column_name(row, "SpanMS", table)/1000)
                            elif insert[1] == "P_NetworkBytesRaw":
                                value = XLSXOutUtils.num_bytes_display(parser.row_element_by_column_name(row, "NetworkBytesRaw", table))
                            elif insert[1] == "P_MBBBytesRaw":
                                value = XLSXOutUtils.num_bytes_display(parser.row_element_by_column_name(row, "MBBBytesRaw", table))
                            else:
                                value = "Could not parse value"
                            out_dict[insert[1]] = XLSXOutUtils.value_to_safe_string(value)

                    # --- Energy Usage {FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}  ---
                    elif table.name == "{FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}":
                        for insert in column_inserts:
                            if insert[1] == "P_AppId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "AppId", table)]
                                #value = parser.row_element_by_column_name(row, "AppId", table)
                            elif insert[1] == "P_UserId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "UserId", table)]
                                #value = parser.row_element_by_column_name(row, "UserId", table)
                            elif insert[1] == "P_EventTimestamp":
                                value = ESEUtils.parse_filetime(parser.row_element_by_column_name(row, "EventTimestamp", table))
                            out_dict[insert[1]] = XLSXOutUtils.value_to_safe_string(value)

                    # --- Long-Term Energy Usage {FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}LT  ---
                    elif table.name == "{FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}LT":
                        for insert in column_inserts:
                            if insert[1] == "P_AppId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "AppId", table)]
                                #value = parser.row_element_by_column_name(row, "AppId", table)
                            elif insert[1] == "P_UserId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "UserId", table)]
                                #value = parser.row_element_by_column_name(row, "UserId", table)
                            elif insert[1] == "P_EventTimestamp":
                                value = ESEUtils.parse_filetime(parser.row_element_by_column_name(row, "EventTimestamp", table))
                            out_dict[insert[1]] = XLSXOutUtils.value_to_safe_string(value)
                    
                    # --- Long-Term Energy Usage {FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}LT  ---
                    elif table.name == "{FEE4E14F-02A9-4550-B5CE-5FA2DA202E37}":
                        for insert in column_inserts:
                            if insert[1] == "P_AppId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "AppId", table)]
                                #value = parser.row_element_by_column_name(row, "AppId", table)
                            elif insert[1] == "P_UserId":
                                value = SruDbIdMap[parser.row_element_by_column_name(row, "UserId", table)]
                                #value = parser.row_element_by_column_name(row, "UserId", table)
                            elif insert[1] == "P_ActiveAcTime":
                                value = ESEUtils.parse_filetime(parser.row_element_by_column_name(row, "ActiveAcTime", table))
                                value = XLSXOutUtils.num_secs_display(parser.row_element_by_column_name(row, "SpanMS", table)/1000)
                            out_dict[insert[1]] = XLSXOutUtils.value_to_safe_string(value)

                    out_row = [out_dict[col_name] for col_name in column_names]
                    worksheet.append(out_row)

            else:
                continue # Not one of the handled sheets


    out_workbook.remove(out_workbook["Sheet"]) # Remove default sheet
    out_workbook.save(output)


export_xlsx()
